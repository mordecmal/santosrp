import csv
from enum import Enum
from typing import List, Tuple, Dict, Set
from datetime import datetime, timedelta
from collections import Counter
import openpyxl
import time
from openpyxl.styles import PatternFill, Font

class DebugStats:
    def __init__(self):
        self.total_citi = 0
        self.total_sidera = 0
        self.total_carriles = 0
        self.matches = Counter()
        self.failed_matches = Counter()
        self.carril_matches = Counter()
        self.camera_stats = {
            'citi': Counter(),
            'sidera': Counter(),
            'carriles': Counter(),
            'common': Counter()
        }
        
    def print_summary(self):
        print("\n=== ESTADÍSTICAS DE COINCIDENCIA ===")
        print(f"Total registros Citi: {self.total_citi}")
        print(f"Total registros Sidera: {self.total_sidera}")
        print(f"Total registros Carriles: {self.total_carriles}")
        
        print("\nTipos de coincidencia:")
        for match_type, count in self.matches.items():
            print(f"  {match_type}: {count}")
        
        print("\nRazones de no coincidencia:")
        reason_translations = {
            'question_mark_camera': 'cámara con signo de interrogación',
            'camera_mismatch': 'cámara no coincide',
            'year_mismatch': 'año no coincide',
            'description_mismatch': 'descripción no coincide',
            'time_mismatch': 'hora fuera del rango permitido',
            'time_parse_error': 'error al procesar hora'
        }
        for reason, count in self.failed_matches.items():
            translated_reason = reason_translations.get(reason, reason)
            print(f"  {translated_reason}: {count}")
        
        print("\nCoincidencias con carriles:")
        for match_type, count in self.carril_matches.items():
            print(f"  {match_type}: {count}")

debug_stats = DebugStats()

def extract_time(time_str: str) -> str:
    """Extract time in HH:MM format"""
    time_str = time_str.strip()
    if ':' in time_str:
        parts = time_str.split(':')
        if len(parts) >= 2:
            return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    return time_str

def is_similar_time(time_str1: str, time_str2: str) -> bool:
    """Compare two time strings with 1-minute tolerance"""
    try:
        time1 = extract_time(time_str1)
        time2 = extract_time(time_str2)
        
        dt1 = datetime.strptime(time1, "%H:%M")
        dt2 = datetime.strptime(time2, "%H:%M")
        
        diff = abs(dt1 - dt2)
        if diff > timedelta(hours=23):
            diff = timedelta(hours=24) - diff
        
        return diff <= timedelta(minutes=1)
    except ValueError as e:
        print(f"Error comparing times: '{time_str1}' and '{time_str2}'")
        raise e

def clean_camera_id(s: str) -> str:
    """Clean camera ID"""
    return s.strip()

def clean_description(s: str) -> str:
    """Clean description"""
    return ' '.join(s.strip().split())

# Constants for empty rows
EMPTY_CITI = ["","","","","",""]
EMPTY_SIDERA = ["","","","","",""]
EMPTY_CARRIL = ["","","",""]

class MatchState(Enum):
    IDENTICAL = 1
    SIMILAR = 2
    DIFFERENT = 3

class Log:
    def __init__(self, line: List, is_citi: bool):
        try:
            self.is_citi = is_citi
            self.raw = line
            self.camera_id = clean_camera_id(line[0])
            
            stats_key = 'citi' if is_citi else 'sidera'
            debug_stats.camera_stats[stats_key][self.camera_id] += 1
            
            if is_citi:
                self.desc = clean_description(line[3])
                self.year = line[4].strip()
                self.hour = line[5].strip()
            else:
                self.desc = clean_description(line[1])
                self.year = line[3].strip()
                self.hour = line[4].strip()
                self.seconds = line[5].strip() if len(line) > 5 else ""
        except IndexError as e:
            print(f"Error processing line: {line}")
            raise e

    def __str__(self):
        return f"{self.camera_id}_{self.year}_{self.hour}"

    def compare(self, other: 'Log', debug: bool = False) -> MatchState:
        if debug:
            print(f"\nComparing logs:")
            print(f"Self:  {self.camera_id} | {self.desc} | {self.year} | {self.hour}")
            print(f"Other: {other.camera_id} | {other.desc} | {other.year} | {other.hour}")

        # If either camera is "?", they can't match
        if self.camera_id == '?' or other.camera_id == '?':
            debug_stats.failed_matches['question_mark_camera'] += 1
            return MatchState.DIFFERENT

        if self.camera_id != other.camera_id:
            debug_stats.failed_matches['camera_mismatch'] += 1
            return MatchState.DIFFERENT

        if self.year != other.year:
            debug_stats.failed_matches['year_mismatch'] += 1
            return MatchState.DIFFERENT

        if self.desc != other.desc:
            debug_stats.failed_matches['description_mismatch'] += 1
            return MatchState.DIFFERENT

        try:
            if self.hour == other.hour:
                return MatchState.IDENTICAL

            if is_similar_time(self.hour, other.hour):
                return MatchState.SIMILAR
            
            debug_stats.failed_matches['time_mismatch'] += 1
        except ValueError:
            debug_stats.failed_matches['time_parse_error'] += 1
            if debug:
                print(f"Time comparison failed for {self} and {other}")
            return MatchState.DIFFERENT
            
        return MatchState.DIFFERENT

class CarrilLog:
    def __init__(self, line: List):
        self.raw = line
        self.camera_prefix = line[0][:6]  # First 6 chars of Equipo
        self.full_id = line[0]
        self.desc = clean_description(line[1])
        self.date = line[2]
        self.hour = line[3].strip()
        
        # Update carril statistics
        debug_stats.camera_stats['carriles'][self.camera_prefix] += 1

    def matches_camera(self, camera_id: str) -> bool:
        """Check if this carril matches a camera ID"""
        return self.camera_prefix == camera_id[:6]

    def matches_time(self, other_hour: str) -> bool:
        """Check if time matches within tolerance"""
        try:
            return is_similar_time(self.hour, other_hour)  # Uses existing 1-minute tolerance
        except ValueError:
            return False

    def matches_event(self, log: Log) -> bool:
        """
        Check if this carril matches a log entry.
        Matches if:
        1. Camera prefix matches
        2. Description is exactly the same
        3. Time is within 1 minute (using existing tolerance)
        """
        # First check the 6-char prefix match
        if not self.matches_camera(log.camera_id):
            return False
            
        # Check description matching
        if self.desc != log.desc:
            return False
            
        # Then check time match with 1-minute tolerance
        if not self.matches_time(log.hour):
            return False
            
        return True

    def __str__(self):
        return f"{self.full_id}_{self.date}_{self.hour}"

class TrafficEvent:
    def __init__(self, log: Log = None):
        self.citi_logs: List[Log] = []
        self.sidera_logs: List[Log] = []
        self.carril_logs: List[CarrilLog] = []
        if log is not None:
            if log.is_citi:
                self.citi_logs.append(log)
            else:
                self.sidera_logs.append(log)

    def add_if_same(self, log: Log) -> bool:
        if log.is_citi:
            if not self.citi_logs:
                self.citi_logs.append(log)
                return True
            if any(existing.compare(log) != MatchState.DIFFERENT for existing in self.citi_logs):
                self.citi_logs.append(log)
                return True
        else:
            if not self.sidera_logs:
                self.sidera_logs.append(log)
                return True
            if any(existing.compare(log) != MatchState.DIFFERENT for existing in self.sidera_logs):
                self.sidera_logs.append(log)
                return True
        return False

    def try_add_carril(self, carril: CarrilLog, used_carriles: Set[CarrilLog]) -> bool:
        """Try to add a carril log if it matches this event"""
        if carril in used_carriles or carril in self.carril_logs:
            return False
            
        matched = False 
        
        for citi_log in self.citi_logs:
            if carril.matches_event(citi_log):
                self.carril_logs.append(carril)
                debug_stats.carril_matches['matched_citi'] += 1
                matched = True
                break  # Stop after first match
                
        # If not matched with Citi, try Sidera logs
        if not matched:
            for sidera_log in self.sidera_logs:
                if carril.matches_event(sidera_log):
                    self.carril_logs.append(carril)
                    debug_stats.carril_matches['matched_sidera'] += 1
                    matched = True
                    break  # Stop after first match

        # For "no coincide" cases
        if not matched and not self.has_match() and (self.citi_logs or self.sidera_logs):
            source_logs = self.citi_logs if self.citi_logs else self.sidera_logs
            for log in source_logs:
                if carril.matches_event(log):
                    self.carril_logs.append(carril)
                    debug_stats.carril_matches['matched_no_coincide'] += 1
                    matched = True
                    break  # Stop after first match

        return matched

    def has_match(self) -> bool:
        """Check if this event has matching Citi and Sidera logs"""
        return len(self.citi_logs) > 0 and len(self.sidera_logs) > 0

    def carril_state(self) -> str:
        """
        Determine the carril state for this event.
        A carril can either be:
        - "incidente con carril" if there are both carriles and incidents
        - "carril sin incidente" if there are only carriles
        - "" (empty string) if there are no carriles
        """
        if len(self.carril_logs) > 0:
            if len(self.citi_logs) > 0 or len(self.sidera_logs) > 0:
                return "incidente con carril"
            return "carril sin incidente"
        return ""  # Empty string for events without carriles

    def title(self) -> str:
        status = self._calculate_title()
        debug_stats.matches[status] += 1
        return status
    
    def _calculate_title(self) -> str:
        if len(self.citi_logs) == 0 and len(self.sidera_logs) == 0:
            if len(self.carril_logs) > 0:
                return "solo carril"
            return ""  # Empty string for completely empty events
        
        # New logic for distinguishing no coincide cases
        if len(self.citi_logs) == 0:
            return "NO COINCIDE SIDERA"  # Only Sidera logs exist
        if len(self.sidera_logs) == 0:
            return "NO COINCIDE CITILOG"  # Only Citi logs exist
    
        if len(self.citi_logs) == 1 and len(self.sidera_logs) == 1:
            match_state = self.citi_logs[0].compare(self.sidera_logs[0])
            if match_state == MatchState.IDENTICAL:
                return "coincide"
            elif match_state == MatchState.SIMILAR:
                return "coincide diff horas"
    
        if len(self.citi_logs) > 1 and len(self.sidera_logs) > 1:
            return "repetido ambos"
        elif len(self.citi_logs) > len(self.sidera_logs):
            return "repetido citi"
        elif len(self.citi_logs) < len(self.sidera_logs):
            return "repetido sidera"
        
        return "no coincide"  # For any other cases that don't match the above conditions

    def has_content(self) -> bool:
        """Check if this event has any actual content"""
        return (len(self.citi_logs) > 0 or 
                len(self.sidera_logs) > 0 or 
                len(self.carril_logs) > 0)

    def return_list(self) -> List:
        if not self.has_content():
            return []  # Return empty list if no actual content
            
        cached_title = self.title()
        max_length = max(len(self.citi_logs), len(self.sidera_logs), len(self.carril_logs))

        sidera_output = [log.raw for log in self.sidera_logs]
        citi_output = [log.raw for log in self.citi_logs]
        carril_output = [log.raw for log in self.carril_logs]

        citi_padded = citi_output + [EMPTY_CITI] * (max_length - len(citi_output))
        sidera_padded = sidera_output + [EMPTY_SIDERA] * (max_length - len(sidera_output))
        carril_padded = carril_output + [EMPTY_CARRIL] * (max_length - len(carril_output))

        return_list = []
        for i in range(max_length):
            citi_row = citi_padded[i]
            sidera_row = sidera_padded[i]
            carril_row = carril_padded[i]
            
            # If this row has only a carril (empty citi and sidera), use empty estado
            if (all(not cell for cell in citi_row) and 
                all(not cell for cell in sidera_row) and 
                any(cell for cell in carril_row)):
                row_title = ""
            else:
                row_title = cached_title

            # Only add rows that have at least some content
            if any(cell for cell in citi_row + sidera_row + carril_row):
                return_list.append(citi_row + sidera_row + carril_row + [row_title, self.carril_state()])

        return return_list

def extract_date_for_sorting(event_row: List) -> datetime:
    try:
        citi_year = event_row[4].strip()
        citi_time = event_row[5].strip()
        if citi_year and citi_time:
            return datetime.strptime(f"{citi_year} {citi_time}", "%Y %H:%M")
            
        sidera_year = event_row[9].strip()
        sidera_time = event_row[10].strip()
        if sidera_year and sidera_time:
            return datetime.strptime(f"{sidera_year} {sidera_time}", "%Y %H:%M")
            
        return datetime.min
    except (ValueError, IndexError):
        return datetime.min

def process_citi_sidera_logs(citi_logs: List[Log], sidera_logs: List[Log], debug: bool = False) -> Tuple[List[TrafficEvent], Set[CarrilLog]]:
    events = []
    used_sidera = set()
    used_citi = set()

    # Group logs by camera ID
    citi_by_camera: Dict[str, List[Tuple[int, Log]]] = {}
    sidera_by_camera: Dict[str, List[Tuple[int, Log]]] = {}

    # First, group all logs by camera ID
    for idx, log in enumerate(citi_logs):
        if log.camera_id not in citi_by_camera:
            citi_by_camera[log.camera_id] = []
        citi_by_camera[log.camera_id].append((idx, log))

    for idx, log in enumerate(sidera_logs):
        if log.camera_id not in sidera_by_camera:
            sidera_by_camera[log.camera_id] = []
        sidera_by_camera[log.camera_id].append((idx, log))

    # Process each camera ID
    for camera_id in set(citi_by_camera.keys()) | set(sidera_by_camera.keys()):
        if camera_id == '?':
            # Handle ? cameras separately
            for citi_idx, citi_log in citi_by_camera.get(camera_id, []):
                if citi_idx not in used_citi:
                    event = TrafficEvent(citi_log)
                    used_citi.add(citi_idx)
                    events.append(event)
            
            for sidera_idx, sidera_log in sidera_by_camera.get(camera_id, []):
                if sidera_idx not in used_sidera:
                    event = TrafficEvent(sidera_log)
                    used_sidera.add(sidera_idx)
                    events.append(event)
            continue

        citi_group = citi_by_camera.get(camera_id, [])
        sidera_group = sidera_by_camera.get(camera_id, [])

        # Try to find matches between Citi and Sidera first
        for citi_idx, citi_log in citi_group:
            if citi_idx in used_citi:
                continue

            best_match = None
            best_match_idx = None
            event = TrafficEvent(citi_log)
            used_citi.add(citi_idx)

            # Look for matching Sidera log
            for sidera_idx, sidera_log in sidera_group:
                if sidera_idx not in used_sidera:
                    match_state = citi_log.compare(sidera_log)
                    if match_state != MatchState.DIFFERENT:
                        best_match = sidera_log
                        best_match_idx = sidera_idx
                        break  # Found a match, no need to continue searching

            # If found a match, add it to the event
            if best_match is not None:
                event.add_if_same(best_match)
                used_sidera.add(best_match_idx)

            # Look for similar Citi logs
            for other_citi_idx, other_citi_log in citi_group:
                if other_citi_idx != citi_idx and other_citi_idx not in used_citi:
                    if citi_log.compare(other_citi_log) != MatchState.DIFFERENT:
                        event.add_if_same(other_citi_log)
                        used_citi.add(other_citi_idx)

            events.append(event)

        # Handle remaining unmatched Sidera logs for this camera
        for sidera_idx, sidera_log in sidera_group:
            if sidera_idx not in used_sidera:
                event = TrafficEvent(sidera_log)
                used_sidera.add(sidera_idx)

                # Look for similar Sidera logs
                for other_sidera_idx, other_sidera_log in sidera_group:
                    if other_sidera_idx != sidera_idx and other_sidera_idx not in used_sidera:
                        if sidera_log.compare(other_sidera_log) != MatchState.DIFFERENT:
                            event.add_if_same(other_sidera_log)
                            used_sidera.add(other_sidera_idx)

                events.append(event)

    return events

def compare_files(citi_path: str, sidera_path: str, carriles_path: str, debug: bool = False, output_path: str = "output.xlsx"):
    start_time = time.time()

def compare_files(citi_path: str, sidera_path: str, carriles_path: str, debug: bool = False, output_path: str = "output.xlsx"):
    print("Iniciando comparación...")
    start_time = time.time()
    debug_stats.__init__()
    
    try:
        # Read input files
        with open(citi_path, encoding='iso-8859-1') as citi_raw:
            citi_log_file = list(csv.reader(citi_raw, delimiter=';'))
        
        with open(sidera_path, encoding='iso-8859-1') as sidera_raw:
            sidera_log_file = list(csv.reader(sidera_raw, delimiter=';'))
            
        with open(carriles_path, encoding='iso-8859-1') as carriles_raw:
            carriles_log_file = list(csv.reader(carriles_raw, delimiter=';'))

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        
        # Create comparison sheet (first sheet)
        comparison_sheet = workbook.active
        comparison_sheet.title = "Comparación"

        # Define styles
        from openpyxl.styles import PatternFill, Font

        # Color definitions with black text for all
        styles = {
            'coincide': {
                'fill': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),  # Green
                'font': Font(color='000000')
            },
            'NO COINCIDE CITILOG': {  # Fixed case to match the status string
                'fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  # Red
                'font': Font(color='000000')
            },
            'coincide diff horas': {
                'fill': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid'),  # Blue
                'font': Font(color='000000')
            },
            'NO COINCIDE SIDERA': {  # Fixed case to match the status string
                'fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  # Red
                'font': Font(color='000000')
            },
            'repetido citi': {
                'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Yellow
                'font': Font(color='000000')
            },
            'repetido sidera': {
                'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Yellow
                'font': Font(color='000000')
            },
            'repetido ambos': {
                'fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),  # Yellow
                'font': Font(color='000000')
            }
}
        # Rest of the code remains the same...

        # Add headers to comparison sheet
        headers = [
            # Citi headers
            "CITILOG CameraName", "CITILOG Start", "CITILOG IncType", 
            "CITILOG IncType", "CITILOG TEXTO AÑOS", "CITILOG TEXTO HORAS",
            # Sidera headers
            "SIDERA Equipo", "SIDERA Desc. variable", "SIDERA Fecha", 
            "SIDERA TEXTO AÑOS", "SIDERA TEXTO HORAS", "SIDERA TEXTO SEGUNDOS",
            # Carril headers
            "CARRIL Equipo", "CARRIL Desc. variable", "CARRIL Fecha", "CARRIL Hora",
            # Status columns
            "ESTADO", "ESTADO CARRIL"
        ]
        comparison_sheet.append(headers)

        # Process data for comparison sheet
        debug_stats.total_citi = len(citi_log_file) - 1
        debug_stats.total_sidera = len(sidera_log_file) - 1
        debug_stats.total_carriles = len(carriles_log_file) - 1

        if debug:
            print(f"Processing {debug_stats.total_citi} Citi logs, {debug_stats.total_sidera} Sidera logs, and {debug_stats.total_carriles} Carril logs")

        # Create Log objects (skip headers)
        citi_logs = [Log(line, True) for line in citi_log_file[1:]]
        sidera_logs = [Log(line, False) for line in sidera_log_file[1:]]
        carril_logs = [CarrilLog(line) for line in carriles_log_file[1:]]

        # Process events
        events = process_citi_sidera_logs(citi_logs, sidera_logs, debug)
        used_carriles = set()

        # Try to match carriles to existing events
        for carril_log in carril_logs:
            if carril_log not in used_carriles:
                matched = False
                for event in events:
                    if event.try_add_carril(carril_log, used_carriles):
                        used_carriles.add(carril_log)
                        matched = True
                if not matched and debug:
                    debug_stats.carril_matches['unmatched'] += 1

        # Create events for unmatched carriles
        for carril_log in carril_logs:
            if carril_log not in used_carriles:
                event = TrafficEvent(None)
                event.carril_logs.append(carril_log)
                events.append(event)
                used_carriles.add(carril_log)
                debug_stats.carril_matches['carril_only'] += 1

        # Add data to comparison sheet
        all_rows = []
        for event in events:
            if event.has_content():
                event_rows = event.return_list()
                all_rows.extend(event_rows)

        # Sort rows by date/time
        sorted_rows = sorted(all_rows, key=extract_date_for_sorting)

        # Filter empty rows
        sorted_rows = [row for row in sorted_rows if any(cell.strip() if isinstance(cell, str) else cell 
                                                       for cell in row[:-2])]

        # Add rows to sheet with formatting
        for row_idx, row in enumerate(sorted_rows, start=2):  # start=2 because row 1 is headers
            comparison_sheet.append(row)
            
            # Get the status (second to last column)
            status = row[-2]
            
            # Apply formatting based on status
            if status in styles:
                for col in range(1, len(row) + 1):  # Excel columns are 1-based
                    cell = comparison_sheet.cell(row=row_idx, column=col)
                    cell.fill = styles[status]['fill']
                    cell.font = styles[status]['font']

        # Create and populate other sheets
        citi_sheet = workbook.create_sheet("Citi")
        for row in citi_log_file:
            citi_sheet.append(row)

        sidera_sheet = workbook.create_sheet("Sidera")
        for row in sidera_log_file:
            sidera_sheet.append(row)

        carriles_sheet = workbook.create_sheet("Carriles")
        for row in carriles_log_file:
            carriles_sheet.append(row)

        # Auto-adjust column widths for all sheets
        for sheet in workbook.sheetnames:
            for column in workbook[sheet].columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                workbook[sheet].column_dimensions[column[0].column_letter].width = adjusted_width

        print(f"\nWriting {output_path}...")
        workbook.save(output_path)
        print(f"Successfully saved {output_path}")
        
        end_time = time.time()
        execution_time = end_time - start_time
        print(f"\nExecution time: {execution_time:.2f} seconds")
        
        # Print statistics
        debug_stats.print_summary()
        
    except Exception as e:
        print(f"Error during comparison: {e}")
        raise e

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Compare Citi, Sidera, and Carriles log files')
    parser.add_argument('citi_path', help='Path to Citi log file')
    parser.add_argument('sidera_path', help='Path to Sidera log file')
    parser.add_argument('carriles_path', help='Path to Carriles log file')
    parser.add_argument('--debug', action='store_true', help='Enable debug output')
    parser.add_argument('--output', default='output.xlsx', help='Output file path (default: output.xlsx)')
    args = parser.parse_args()
    
    compare_files(args.citi_path, args.sidera_path, args.carriles_path, args.debug, args.output)
